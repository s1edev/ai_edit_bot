"""Admin handlers."""

from aiogram import Router
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, CallbackQuery, InputMediaPhoto, InputMediaVideo, InputMediaAnimation
from aiogram.types import FSInputFile
from utils.helpers import process_user_payment
from db.database import AsyncSessionLocal
from db.requests_db import UserRepository
from filters.filters import AdminFilter
from lexicon.lexicon import ADMIN_LEXICON
from fsm.fsm import AdminState
from keyboards.keyboards import get_newsletter_buttons_keyboard, get_newsletter_preview_keyboard, get_newsletter_media_type_keyboard
import openpyxl
from openpyxl.styles import Font, Alignment
import tempfile
import os

router = Router(name="admin_router")

@router.message(Command("admin_stats"), AdminFilter())
async def admin_stats(message: Message) -> None:
    async with AsyncSessionLocal() as session:
        users_count = await UserRepository.get_users_count(session)
        if users_count == 0:
            await message.answer(ADMIN_LEXICON["admin_stats_empty"])
            return
        all_users = await UserRepository.get_all_users(session)
        total_coins = sum(u.coins for u in all_users)
        avg_coins = total_coins / users_count
        text = (
            ADMIN_LEXICON["admin_stats_header"] +
            ADMIN_LEXICON["admin_stats_users"].format(users_count=users_count) +
            ADMIN_LEXICON["admin_stats_total_coins"].format(total_coins=total_coins) +
            ADMIN_LEXICON["admin_stats_avg_coins"].format(avg_coins=avg_coins)
        )
        await message.answer(text, parse_mode="Markdown")


@router.message(Command("admin_help"), AdminFilter())
async def admin_help(message: Message) -> None:
    text = (
        ADMIN_LEXICON["admin_help_header"] +
        ADMIN_LEXICON["admin_help_stats"] +
        ADMIN_LEXICON["admin_help_newsletter"] +
        ADMIN_LEXICON["admin_help_add_coins"] +
        ADMIN_LEXICON["admin_help_subtract_coins"] +
        ADMIN_LEXICON["admin_help_subtract_money"] +
        ADMIN_LEXICON["admin_help_add_to_user_fsm"] +
        ADMIN_LEXICON["admin_help_set_referral_percentage"] +
        ADMIN_LEXICON["admin_help_help"]
    )
    await message.answer(text)

@router.message(Command("add_coins"), AdminFilter())
async def add_coins(message: Message) -> None:
    try:
        args = message.text.split()[1:]
        if len(args) < 2:
            await message.answer(ADMIN_LEXICON["add_coins_usage"])
            return
        tg_id, amount = int(args[0]), int(args[1])
        async with AsyncSessionLocal() as session:
            user = await UserRepository.get_user_by_tg_id(session, tg_id)
            if not user:
                await message.answer(ADMIN_LEXICON["user_not_found"].format(tg_id=tg_id))
                return
            u = await UserRepository.add_coins(session, user.id, amount)
            if u:
                await message.answer(ADMIN_LEXICON["add_coins_success"].format(tg_id=tg_id, amount=amount, new_balance=u.coins))
            else:
                await message.answer(ADMIN_LEXICON["operation_failed"])
    except (ValueError, IndexError):
        await message.answer(ADMIN_LEXICON["add_coins_usage"])


@router.message(Command("cancel"), StateFilter(AdminState))
async def cancel_admin_fsm(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(ADMIN_LEXICON["add_to_user_fsm_cancelled"])


@router.message(Command("add_to_user_fsm"), AdminFilter())
async def add_to_user_fsm_start(message: Message, state: FSMContext) -> None:
    await message.answer(ADMIN_LEXICON["add_to_user_fsm_start"])
    await state.set_state(AdminState.add_to_user_tg_id)


@router.message(AdminState.add_to_user_tg_id, AdminFilter())
async def add_to_user_fsm_tg_id(message: Message, state: FSMContext) -> None:
    try:
        await state.update_data(tg_id=int(message.text))
        await message.answer(ADMIN_LEXICON["add_to_user_fsm_tg_id"])
        await state.set_state(AdminState.add_to_user_coins)
    except ValueError:
        await message.answer("❌ Неверный формат Telegram ID.")


@router.message(AdminState.add_to_user_coins, AdminFilter())
async def add_to_user_fsm_coins(message: Message, state: FSMContext) -> None:
    try:
        await state.update_data(coins=int(message.text))
        await message.answer(ADMIN_LEXICON["add_to_user_fsm_money"])
        await state.set_state(AdminState.add_to_user_money)
    except ValueError:
        await message.answer("❌ Неверный формат количества монет.")


@router.message(AdminState.add_to_user_money, AdminFilter())
async def add_to_user_fsm_money(message: Message, state: FSMContext) -> None:
    try:
        await state.update_data(money=int(message.text))
        await message.answer(ADMIN_LEXICON["add_to_user_fsm_description"])
        await state.set_state(AdminState.add_to_user_description)
    except ValueError:
        await message.answer("❌ Неверный формат суммы платежа.")


@router.message(AdminState.add_to_user_description, AdminFilter())
async def add_to_user_fsm_description(message: Message, state: FSMContext) -> None:
    from utils.helpers import process_user_payment
    description = message.text if message.text.lower() != "пропустить" else ""
    await state.update_data(description=description)
    data = await state.get_data()
    tg_id, coins, money = data["tg_id"], data["coins"], data.get("money", 0)
    if coins == 0:
        await message.answer(ADMIN_LEXICON["add_coins_usage"])
        await state.clear()
        return
    async with AsyncSessionLocal() as session:
        user = await UserRepository.get_user_by_tg_id(session, tg_id)
        if not user:
            await message.answer(ADMIN_LEXICON["user_not_found"].format(tg_id=tg_id))
            await state.clear()
            return
        operations = await process_user_payment(session, tg_id, coins, money, description)
        await message.answer("\n".join(operations) if operations else ADMIN_LEXICON["operation_failed"])
    await state.clear()
    await message.answer(ADMIN_LEXICON["add_to_user_fsm_completed"])


@router.message(Command("set_referral_percentage"), AdminFilter())
async def set_referral_percentage(message: Message) -> None:
    try:
        args = message.text.split()[1:]
        if len(args) != 2:
            await message.answer(ADMIN_LEXICON["set_referral_percentage_usage"])
            return
        user_identifier, percentage = args[0], int(args[1])
        if not 0 <= percentage <= 100:
            await message.answer("❌ Процент 0-100")
            return
        async with AsyncSessionLocal() as session:
            user = await UserRepository.get_user_by_tg_id(session, int(user_identifier)) if user_identifier.isdigit() else await UserRepository.get_user_by_hash(session, user_identifier)
            if not user:
                await message.answer(ADMIN_LEXICON["user_not_found"].format(tg_id=user_identifier))
                return
            old = user.referral_percentage
            u = await UserRepository.update_referral_percentage(session, user.id, percentage)
            if u:
                await message.answer(ADMIN_LEXICON["set_referral_percentage_success"].format(tg_id=user_identifier, percentage=u.referral_percentage, old_percentage=old))
            else:
                await message.answer(ADMIN_LEXICON["operation_failed"])
    except (ValueError, IndexError):
        await message.answer(ADMIN_LEXICON["set_referral_percentage_usage"])


@router.message(Command("subtract_coins"), AdminFilter())
async def subtract_coins(message: Message) -> None:
    try:
        args = message.text.split()[1:]
        if len(args) < 2:
            await message.answer(ADMIN_LEXICON["subtract_coins_usage"])
            return
        tg_id, amount = int(args[0]), int(args[1])
        async with AsyncSessionLocal() as session:
            user = await UserRepository.get_user_by_tg_id(session, tg_id)
            if not user:
                await message.answer(ADMIN_LEXICON["user_not_found"].format(tg_id=tg_id))
                return
            u = await UserRepository.subtract_coins(session, user.id, amount)
            if u:
                await message.answer(ADMIN_LEXICON["subtract_coins_success"].format(tg_id=tg_id, amount=amount, new_balance=u.coins))
            else:
                await message.answer(ADMIN_LEXICON["operation_failed"])
    except (ValueError, IndexError):
        await message.answer(ADMIN_LEXICON["subtract_coins_usage"])


@router.message(Command("subtract_money"), AdminFilter())
async def subtract_money(message: Message) -> None:
    try:
        args = message.text.split()[1:]
        if len(args) < 2:
            await message.answer(ADMIN_LEXICON["subtract_money_usage"])
            return
        tg_id, amount = int(args[0]), int(args[1])
        async with AsyncSessionLocal() as session:
            user = await UserRepository.get_user_by_tg_id(session, tg_id)
            if not user:
                await message.answer(ADMIN_LEXICON["user_not_found"].format(tg_id=tg_id))
                return
            u = await UserRepository.subtract_referral_earnings(session, user.id, amount)
            if u:
                await message.answer(ADMIN_LEXICON["subtract_money_success"].format(tg_id=tg_id, amount=amount, new_earnings=u.referral_earnings))
            else:
                await message.answer(ADMIN_LEXICON["operation_failed"])
    except (ValueError, IndexError):
        await message.answer(ADMIN_LEXICON["subtract_money_usage"])


@router.message(Command("newsletter"), AdminFilter())
async def newsletter_start(message: Message, state: FSMContext) -> None:
    """Начало создания рассылки."""
    await message.answer(ADMIN_LEXICON["newsletter_start"], reply_markup=get_newsletter_media_type_keyboard())
    await state.set_state(AdminState.newsletter_media_type)


# Callback handlers for newsletter creation
@router.callback_query(lambda c: c.data.startswith("media_type_"))
async def newsletter_media_type_selected(callback: CallbackQuery, state: FSMContext):
    """Обработка выбора типа медиа."""
    media_type = callback.data.replace("media_type_", "")
    media_names = {
        "photo": "📷 Фото",
        "video": "🎥 Видео", 
        "gif": "🎞️ GIF",
        "text": "📝 Только текст"
    }
    
    await state.update_data(media_type=media_type)
    
    if media_type == "text":
        await callback.message.answer(ADMIN_LEXICON["newsletter_text_prompt"])
        await state.set_state(AdminState.newsletter_text)
    else:
        await callback.message.answer(
            ADMIN_LEXICON["newsletter_media_type_selected"].format(media_type=media_names[media_type])
        )
        await state.set_state(AdminState.newsletter_media_content)
    
    await callback.answer()


@router.callback_query(lambda c: c.data == "cancel_newsletter")
async def cancel_newsletter(callback: CallbackQuery, state: FSMContext):
    """Отмена создания рассылки."""
    await state.clear()
    await callback.message.answer(ADMIN_LEXICON["newsletter_cancelled"])
    await callback.answer()


@router.callback_query(lambda c: c.data == "add_button")
async def newsletter_add_button(callback: CallbackQuery, state: FSMContext):
    """Добавление кнопки к рассылке."""
    await callback.message.answer(ADMIN_LEXICON["newsletter_button_text_prompt"])
    await state.set_state(AdminState.newsletter_button_text)
    await callback.answer()


@router.callback_query(lambda c: c.data == "skip_button")
async def newsletter_skip_button(callback: CallbackQuery, state: FSMContext):
    """Пропуск добавления кнопки."""
    await state.update_data(button_text=None, button_url=None)
    await send_newsletter_preview(callback.message, state)
    await callback.answer()


@router.callback_query(lambda c: c.data == "send_newsletter")
async def newsletter_send(callback: CallbackQuery, state: FSMContext):
    """Отправка рассылки."""
    await send_newsletter_to_users(callback.message, state)
    await callback.answer()


@router.callback_query(lambda c: c.data == "edit_newsletter")
async def newsletter_edit(callback: CallbackQuery, state: FSMContext):
    """Редактирование рассылки."""
    await state.clear()
    await newsletter_start(callback.message, state)
    await callback.answer()


# Message handlers for newsletter creation
@router.message(AdminState.newsletter_media_content, AdminFilter())
async def newsletter_media_content(message: Message, state: FSMContext):
    """Получение медиафайла для рассылки."""
    data = await state.get_data()
    media_type = data.get("media_type")
    
    if media_type == "photo" and message.photo:
        file_id = message.photo[-1].file_id
        await state.update_data(media_file_id=file_id)
    elif media_type == "video" and message.video:
        file_id = message.video.file_id
        await state.update_data(media_file_id=file_id)
    elif media_type == "gif" and message.animation:
        file_id = message.animation.file_id
        await state.update_data(media_file_id=file_id)
    else:
        await message.answer("❌ Пожалуйста, отправьте файл соответствующего типа.")
        return
    
    await message.answer(ADMIN_LEXICON["newsletter_text_prompt"])
    await state.set_state(AdminState.newsletter_text)


@router.message(AdminState.newsletter_text, AdminFilter())
async def newsletter_text(message: Message, state: FSMContext):
    """Получение текста для рассылки."""
    await state.update_data(text=message.text)
    await message.answer(ADMIN_LEXICON["newsletter_button_prompt"], reply_markup=get_newsletter_buttons_keyboard())


@router.message(AdminState.newsletter_button_text, AdminFilter())
async def newsletter_button_text(message: Message, state: FSMContext):
    """Получение текста кнопки."""
    await state.update_data(button_text=message.text)
    await message.answer(ADMIN_LEXICON["newsletter_button_url_prompt"])
    await state.set_state(AdminState.newsletter_button_url)


@router.message(AdminState.newsletter_button_url, AdminFilter())
async def newsletter_button_url(message: Message, state: FSMContext):
    """Получение URL кнопки."""
    await state.update_data(button_url=message.text)
    await send_newsletter_preview(message, state)


async def send_newsletter_preview(message: Message, state: FSMContext):
    """Отправка предварительного просмотра рассылки."""
    data = await state.get_data()
    media_type = data.get("media_type")
    text = data.get("text")
    button_text = data.get("button_text")
    button_url = data.get("button_url")
    
    # Создаем клавиатуру с кнопкой если она есть
    keyboard = None
    if button_text and button_url:
        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text=button_text, url=button_url)]
            ]
        )
    
    # Отправляем предварительный просмотр
    if media_type == "photo" and data.get("media_file_id"):
        await message.bot.send_photo(
            chat_id=message.chat.id,
            photo=data["media_file_id"],
            caption=text,
            reply_markup=keyboard
        )
    elif media_type == "video" and data.get("media_file_id"):
        await message.bot.send_video(
            chat_id=message.chat.id,
            video=data["media_file_id"],
            caption=text,
            reply_markup=keyboard
        )
    elif media_type == "gif" and data.get("media_file_id"):
        await message.bot.send_animation(
            chat_id=message.chat.id,
            animation=data["media_file_id"],
            caption=text,
            reply_markup=keyboard
        )
    else:
        await message.answer(text, reply_markup=keyboard)
    
    # Получаем количество пользователей
    async with AsyncSessionLocal() as session:
        users_count = await UserRepository.get_users_count(session)
    
    if users_count == 0:
        await message.answer(ADMIN_LEXICON["newsletter_no_users"])
        await state.clear()
        return
    
    await message.answer(
        ADMIN_LEXICON["newsletter_confirm_send"].format(count=users_count),
        reply_markup=get_newsletter_preview_keyboard()
    )


async def send_newsletter_to_users(message: Message, state: FSMContext):
    """Отправка рассылки всем пользователям."""
    data = await state.get_data()
    media_type = data.get("media_type")
    text = data.get("text")
    button_text = data.get("button_text")
    button_url = data.get("button_url")
    
    # Создаем клавиатуру с кнопкой если она есть
    keyboard = None
    if button_text and button_url:
        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text=button_text, url=button_url)]
            ]
        )
    
    # Получаем список всех пользователей
    async with AsyncSessionLocal() as session:
        users = await UserRepository.get_all_users(session)
    
    if not users:
        await message.answer(ADMIN_LEXICON["newsletter_no_users"])
        await state.clear()
        return
    
    # Отправляем рассылку
    sent_count = 0
    total_count = len(users)
    
    await message.answer(ADMIN_LEXICON["newsletter_sending"].format(sent=sent_count, total=total_count))
    
    for user in users:
        try:
            if media_type == "photo" and data.get("media_file_id"):
                await message.bot.send_photo(
                    chat_id=user.tg_id,
                    photo=data["media_file_id"],
                    caption=text,
                    reply_markup=keyboard
                )
            elif media_type == "video" and data.get("media_file_id"):
                await message.bot.send_video(
                    chat_id=user.tg_id,
                    video=data["media_file_id"],
                    caption=text,
                    reply_markup=keyboard
                )
            elif media_type == "gif" and data.get("media_file_id"):
                await message.bot.send_animation(
                    chat_id=user.tg_id,
                    animation=data["media_file_id"],
                    caption=text,
                    reply_markup=keyboard
                )
            else:
                await message.bot.send_message(
                    chat_id=user.tg_id,
                    text=text,
                    reply_markup=keyboard
                )
            
            sent_count += 1
            
            # Обновляем сообщение о прогрессе каждые 10 сообщений
            if sent_count % 10 == 0:
                await message.edit_text(ADMIN_LEXICON["newsletter_sending"].format(sent=sent_count, total=total_count))
                
        except Exception as e:
            # Игнорируем ошибки (пользователь мог заблокировать бота)
            pass
    
    await message.edit_text(ADMIN_LEXICON["newsletter_completed"].format(sent=sent_count))
    await state.clear()


@router.message(Command("download_users_db"), AdminFilter())
async def download_users_db(message: Message) -> None:
    """Download users database as Excel file and bot.db."""
    await message.answer("⏳ Подготавливаю базу данных пользователей...")
    
    try:
        async with AsyncSessionLocal() as session:
            users = await UserRepository.get_all_users(session)
            
            if not users:
                await message.answer("❌ База данных пуста")
                return
            
            # Create Excel workbook with multiple sheets
            wb = openpyxl.Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Sheet 1: Users
            ws_users = wb.create_sheet(title="Пользователи")
            users_headers = [
                "ID", "Telegram ID", "Username", "Хеш", "Приглашен", "Монеты", 
                "Заработок", "Процент", "Приглашено", "Дата регистрации"
            ]
            
            for col, header in enumerate(users_headers, 1):
                cell = ws_users.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            for row, user in enumerate(users, 2):
                ws_users.cell(row=row, column=1, value=user.id)
                ws_users.cell(row=row, column=2, value=user.tg_id)
                ws_users.cell(row=row, column=3, value=user.username or "")
                ws_users.cell(row=row, column=4, value=user.user_hash)
                ws_users.cell(row=row, column=5, value=user.invited_by_hash or "")
                ws_users.cell(row=row, column=6, value=user.coins)
                ws_users.cell(row=row, column=7, value=user.referral_earnings)
                ws_users.cell(row=row, column=8, value=user.referral_percentage)
                ws_users.cell(row=row, column=9, value=user.invited_count)
                ws_users.cell(row=row, column=10, value=user.registered_at.strftime("%Y-%m-%d %H:%M:%S"))
            
            # Sheet 2: Transactions
            from db.requests_db import TransactionRepository
            transactions = await TransactionRepository.get_all_transactions(session)
            ws_transactions = wb.create_sheet(title="Транзакции")
            transactions_headers = [
                "ID", "User ID", "Telegram ID", "Монеты", "Сумма", "Описание", "Дата"
            ]
            
            for col, header in enumerate(transactions_headers, 1):
                cell = ws_transactions.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            for row, transaction in enumerate(transactions, 2):
                ws_transactions.cell(row=row, column=1, value=transaction.id)
                ws_transactions.cell(row=row, column=2, value=transaction.user_id)
                ws_transactions.cell(row=row, column=3, value=transaction.tg_id)
                ws_transactions.cell(row=row, column=4, value=transaction.coins)
                ws_transactions.cell(row=row, column=5, value=transaction.money)
                ws_transactions.cell(row=row, column=6, value=transaction.description or "")
                ws_transactions.cell(row=row, column=7, value=transaction.created_at.strftime("%Y-%m-%d %H:%M:%S"))
            
            # Sheet 3: Usage
            from db.requests_db import UsageRepository
            usage_records = await UsageRepository.get_all_usage(session)
            ws_usage = wb.create_sheet(title="Использование")
            usage_headers = [
                "ID", "User ID", "Telegram ID", "Монеты", "Дата"
            ]
            
            for col, header in enumerate(usage_headers, 1):
                cell = ws_usage.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            for row, usage in enumerate(usage_records, 2):
                ws_usage.cell(row=row, column=1, value=usage.id)
                ws_usage.cell(row=row, column=2, value=usage.user_id)
                # Usage model doesn't have tg_id, only user_id. 
                # To get tg_id, we would need to join with User table.
                # For now, we'll leave this column empty or use user_id as reference.
                ws_usage.cell(row=row, column=3, value="")  # tg_id not available in Usage model
                ws_usage.cell(row=row, column=4, value=usage.coins_used)
                ws_usage.cell(row=row, column=5, value=usage.used_at.strftime("%Y-%m-%d %H:%M:%S"))
            
            # Auto-adjust column widths for all sheets
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_filename = temp_file.name
                wb.save(temp_filename)
            
            try:
                # Check Excel file size before sending
                excel_size = os.path.getsize(temp_filename)
                if excel_size > 50 * 1024 * 1024:  # 50MB limit
                    await message.answer("❌ Excel файл слишком большой для отправки (>50MB)")
                else:
                    # Send Excel file
                    excel_file = FSInputFile(
                        path=temp_filename,
                        filename="users_database.xlsx"
                    )
                    
                    await message.answer_document(
                        document=excel_file,
                        caption=f"📊 База данных пользователей\n👥 Всего пользователей: {len(users)}\n💰 Транзакций: {len(transactions)}\n🔄 Использований: {len(usage_records)}"
                    )
                
                # Send bot.db file with size check
                db_path = "data/bot.db"
                if os.path.exists(db_path):
                    db_size = os.path.getsize(db_path)
                    if db_size > 50 * 1024 * 1024:  # 50MB limit
                        await message.answer("❌ Файл bot.db слишком большой для отправки (>50MB)")
                    else:
                        db_file = FSInputFile(
                            path=db_path,
                            filename="bot.db"
                        )
                        
                        await message.answer_document(
                            document=db_file,
                            caption="🗄️ bot.db"
                        )
                else:
                    await message.answer("⚠️ Файл bot.db не найден в папке data/")
                
            except Exception as e:
                await message.answer(f"❌ Ошибка при отправке файлов: {str(e)}")
            finally:
                # Clean up temporary file
                if os.path.exists(temp_filename):
                    os.unlink(temp_filename)
            
    except Exception as e:
        await message.answer(f"❌ Ошибка при создании файла: {str(e)}")
